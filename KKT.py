import openpyxl
import webbrowser
import selenium
from selenium import webdriver
loadedsheet=openpyxl.load_workbook("PR.xlsx")#载入PR表
sheet=loadedsheet.get_sheet_by_name('Sheet1') #选中PR清单中的sheet1
rowsnumber=sheet.max_row+1 #求PR的总行数

for i in range(rowsnumber):  #以下是一个“有多少行PR就使用多少次”的循环
    if i==0:
        continue
    PR=sheet.cell(row=i,column=1).value  #读取清单表格中第i行，第1列（PR所在的列数），即PR所在位置 
    if PR==None:
        break
    webbrowser.open(f"https://jabil.coupahost.com/approver/{PR}/edit") #打开PR的对应的核单网页
    #driver= webdriver.Edge()
    #driver.find_element_by_xpath('//*[@id="requisition_header_custom_field_3"]').clear()#清空输入框中已有的值
    #driver.find_element_by_xpath('//*[@id="requisition_header_custom_field_3"]').send_keys("donald tang")
